import pandas as pd
import numpy as np
import time,datetime
from datetime import datetime, timedelta
import yaml
import os
from openpyxl import load_workbook
# folder to load config file
CONFIG_PATH = "./Data initializers/"
# Function to load yaml configuration file
def load_config(config_name):
    with open(os.path.join(CONFIG_PATH, config_name)) as file:
        config = yaml.safe_load(file)
    return config

config = load_config("data_loader_results.yaml")
type_description = pd.read_excel(os.path.join(config["data_directory"], config["data_name"]),sheet_name=config['sheet_type_description'])

type_list = sorted(type_description[config['type']].unique())
description_index = pd.DataFrame(columns=['TYPE', 'DESCRIPTION'])
for type in type_list:
    data_type = type_description[type_description[config['type']] == type]
    description_list = sorted(data_type[config['description']].unique())
    for description in description_list:
        description_index = description_index.append({'TYPE': type, 'DESCRIPTION': description}, ignore_index=True)

from openpyxl import load_workbook

path_result = os.path.join(config["data_directory"], config["data_name"])
book = load_workbook(path_result)
writer_result = pd.ExcelWriter(path_result, engine='openpyxl')
writer_result.book = book
description_index.to_excel(writer_result, sheet_name=config['sheet_index_description'])
writer_result.save()

path_result = r"C:\\Users\\User\\Desktop\\B-Yond\\New_paper-Alarms\\Approach3-Sample2\\Results.xlsx"
book = load_workbook(path_result)
writer_result = pd.ExcelWriter(path_result, engine='openpyxl')
writer_result.book = book
description_index.to_excel(writer_result, sheet_name='Description_index')
writer_result.save()
