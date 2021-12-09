import pandas as pd
import numpy as np
import time,datetime
from datetime import datetime, timedelta
import yaml
import os
from scipy.stats import poisson
from scipy.stats import expon
import math
from openpyxl import load_workbook

# folder to load config file
CONFIG_PATH = "./Data initializers/"
# Function to load yaml configuration file
def load_config(config_name):
    with open(os.path.join(CONFIG_PATH, config_name)) as file:
        config = yaml.safe_load(file)
    return config

config = load_config("data_loader_results.yaml")
number_of_alarms = pd.read_excel(os.path.join(config["data_directory"], config["data_name"]),sheet_name=config['sheet_score_type'])
type_list = sorted(number_of_alarms[config['type']].unique())

max_min_type = pd.DataFrame(columns=['TYPE', 'Min_P2', 'Max_P2', 'Min_P3', 'Max_P3', 'Min_S4', 'Max_S4'])
for type in type_list:
    number_of_alarms_type = number_of_alarms[number_of_alarms[config['type']] == type]
    Max_S4 = number_of_alarms_type[config['score_frequency']].max()
    Min_S4 = number_of_alarms_type[config['score_frequency']].min()
    number_of_alarms_type_greater1 = number_of_alarms_type[number_of_alarms_type[config['number_of_alarms']] != 1]
    Max_P2 = number_of_alarms_type_greater1[config['exponential_probability']].max()
    Max_P3 = number_of_alarms_type_greater1[config['transition_probability']].max()
    Min_P2 = number_of_alarms_type_greater1[config['exponential_probability']].min()
    Min_P3 = number_of_alarms_type_greater1[config['transition_probability']].min()
    max_min_type = max_min_type.append(
        {'TYPE': type, 'Min_P2': Min_P2, 'Max_P2': Max_P2, 'Min_P3': Min_P3, 'Max_P3': Max_P3, 'Min_S4': Min_S4,'Max_S4': Max_S4}, ignore_index=True)

from openpyxl import load_workbook
path_result = os.path.join(config["data_directory"], config["data_name"])
book = load_workbook(path_result)
writer_result = pd.ExcelWriter(path_result, engine='openpyxl')
writer_result.book = book

max_min_type.to_excel(writer_result, sheet_name=config['sheet_max_min'])
writer_result.save()


