import pandas as pd
import numpy as np
import time,datetime
from datetime import datetime, timedelta
import yaml
import os
from openpyxl import load_workbook
# folder to load config file
CONFIG_PATH = "./Data initializers/"
# Function to load yaml configuration file
def load_config(config_name):
    with open(os.path.join(CONFIG_PATH, config_name)) as file:
        config = yaml.safe_load(file)
    return config

config = load_config("data_loader_aggregated.yaml")
# load data
aggregated_alarms = pd.read_csv(os.path.join(config["data_directory"], config["data_name"]))
aggregated_alarms=aggregated_alarms.iloc[:,1:]

S4 = pd.DataFrame(columns=['TYPE', 'DESCRIPTION', 'FREQUENCY', 'S4'])
type_list = sorted(aggregated_alarms[config['type']].unique())

for type in type_list:
    aggregated_alarms_type = aggregated_alarms[aggregated_alarms[config['type']] == type]
    description_list = sorted(aggregated_alarms_type[config['description']].unique())
    for description in description_list:
        aggregated_alarms_type_description = aggregated_alarms_type[aggregated_alarms_type[config['description']] == description]
        S4 = S4.append({'TYPE': type, 'DESCRIPTION': description, 'FREQUENCY': len(aggregated_alarms_type_description), 'S4': 1 / len(aggregated_alarms_type_description)}, ignore_index=True)

config_result = load_config("data_loader_results.yaml")

from openpyxl import load_workbook

path_result = os.path.join(config_result["data_directory"], config_result["data_name"])
book = load_workbook(path_result)
writer_result = pd.ExcelWriter(path_result, engine='openpyxl')
writer_result.book = book

S4.to_excel(writer_result, sheet_name=config_result['sheet_S4'])
writer_result.save()

